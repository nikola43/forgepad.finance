#!/usr/bin/env python3
"""Build the Fyuz graduation report workbook from GraduationCurveTest output.

Usage:
    python3 scripts/graduation_report.py [forge-log-file]

With no argument, runs the capture test itself:
    forge test --match-contract GraduationCurveTest -vv   (in foundry/)

Parses "GCURVE,<market>,<phase>,<step>,<bnbIn_wei>,<priceBNB_wei>,<bnbUsd_wei>,
<mcapUsd_wei>" rows and writes docs/reports/fyuz_graduation_report.xlsx —
one sheet with summary, both data tables (V2 + V3) and charts.

Prices are stored as real values (e.g. 0.0000000077 BNB) with fixed-decimal
number formats. After saving, the worksheet XML is rewritten to plain-decimal
notation — some spreadsheet apps mis-parse scientific notation like 7.7e-09.

Requires: openpyxl (pip install openpyxl).
"""
import re
import subprocess
import sys
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "docs/reports/fyuz_graduation_report.xlsx"

# CVD-safe pair validated with the dataviz palette checker.
BLUE = "2A78D6"    # bonding curve
ORANGE = "E0662B"  # DEX opening price
GREEN = "008300"   # market cap
GRAY = "9CA3AF"    # graduation target reference line

INK = "1F2937"     # header fill
BAND = "F3F4F6"    # zebra band

BNB_FMT = "0.0000000000000"  # 13 dp — BNB prices are ~1e-8
USD_FMT = "0.0000000000"     # 10 dp — USD prices are ~1e-5


def capture_rows(log_text):
    """-> {market: [(step, bnb_in, price_bnb, price_usd, mcap_usd, phase)]}"""
    data = {}
    for line in log_text.splitlines():
        line = line.strip()
        i = line.find("GCURVE,")
        if i < 0:
            continue
        _, market, phase, step, bnb_in, price_bnb, bnb_usd, mcap = line[i:].split(",")
        price_bnb = int(price_bnb) / 1e18
        bnb_usd = int(bnb_usd) / 1e18
        data.setdefault(market, []).append((
            int(step),
            int(bnb_in) / 1e18,
            price_bnb,
            price_bnb * bnb_usd,
            int(mcap) / 1e18,
            phase,
        ))
    for market, rows in data.items():
        assert rows, f"no rows for {market}"
        curve = [r for r in rows if r[5] == "curve"]
        assert all(a[2] <= b[2] for a, b in zip(curve, curve[1:])), \
            f"{market}: curve price not monotonic"
    return data


def styled_header(ws, row, cols, texts):
    fill = PatternFill("solid", fgColor=INK)
    font = Font(bold=True, color="FFFFFF", size=10)
    for col, text in zip(cols, texts):
        c = ws.cell(row=row, column=col, value=text)
        c.fill, c.font = fill, font
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")


def line_series(ws, min_col, min_row, max_row, title, color, width=28000,
                marker=False, dashed=False):
    s = Series(
        Reference(ws, min_col=min_col, min_row=min_row, max_row=max_row),
        title_from_data=False,
        title=title,
    )
    lp = LineProperties(solidFill=color, w=width)
    if dashed:
        lp.prstDash = "dash"
    s.graphicalProperties.line = lp
    s.marker = Marker(symbol="circle", size=7) if marker else Marker(symbol="none")
    if marker:
        s.marker.graphicalProperties.solidFill = color
        s.marker.graphicalProperties.line.solidFill = color
    return s


def make_chart(title, y_title):
    ch = LineChart()
    ch.title = title
    ch.style = None
    ch.y_axis.title = y_title
    ch.x_axis.title = "Cumulative BNB (curve raise + DEX volume)"
    ch.x_axis.delete = ch.y_axis.delete = False
    ch.width, ch.height = 21, 10
    ch.x_axis.tickLblSkip = 3
    ch.legend.overlay = False
    return ch


def build(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Graduation Report"
    ws.sheet_view.showGridLines = False

    # ---- Title block -------------------------------------------------------
    ws.merge_cells("B2:K2")
    t = ws["B2"]
    t.value = "Fyuz Graduation Report — bonding curve → DEX (BSC mainnet fork)"
    t.font = Font(bold=True, size=16, color=INK)
    ws.merge_cells("B3:K3")
    v2, v3 = data["v2"], data["v3"]
    bnb_usd = v2[0][3] / v2[0][2]  # USD price / BNB price at step 0
    sub = ws["B3"]
    sub.value = (f"Real PancakeSwap + Chainlink on a BNB Chain fork · "
                 f"BNB ≈ ${bnb_usd:,.0f} at capture")
    sub.font = Font(size=10, color="6B7280", italic=True)

    # ---- Summary -----------------------------------------------------------
    def summary(rows):
        curve = [r for r in rows if r[5] == "curve"]
        dex = [r for r in rows if r[5] == "dex"]
        gap = (dex[0][2] / curve[-1][2] - 1) * 100
        dex_mcap = curve[-1][4] * dex[0][2] / curve[-1][2]
        return {
            "Opening market cap (USD)": f"${curve[0][4]:,.0f}",
            "Market cap at DEX open (USD)": f"${dex_mcap:,.0f}",
            "BNB raised to graduate": f"{dex[0][1]:,.1f} BNB",
            "Opening price": f"{curve[0][2]:.12f} BNB",
            "Last curve price": f"{curve[-1][2]:.12f} BNB",
            "DEX opening price": f"{dex[0][2]:.12f} BNB",
            "Curve → DEX price gap": f"{gap:+.2f}%",
            f"DEX price after {len(dex) - 1} trades": f"{dex[-1][2]:.12f} BNB",
            "Price appreciation to DEX": f"{dex[0][2] / curve[0][2]:.1f}×",
        }

    s2, s3 = summary(v2), summary(v3)
    styled_header(ws, 5, [2, 3, 4], ["Metric", "PancakeSwap V2", "PancakeSwap V3"])
    thin = Side(style="thin", color="D1D5DB")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, k in enumerate(s2):
        r = 6 + i
        for col, val in ((2, k), (3, s2[k]), (4, s3[k])):
            c = ws.cell(row=r, column=col, value=val)
            c.border = box
            c.font = Font(size=10, bold=(col == 2))
            if i % 2:
                c.fill = PatternFill("solid", fgColor=BAND)
            if col > 2:
                c.alignment = Alignment(horizontal="right")

    # ---- Data tables (V2 at F, V3 at N) ------------------------------------
    HEADERS = ["Step", "Phase", "BNB raised",
               "Curve price\n(BNB)", "DEX price\n(BNB)",
               "Curve price\n(USD)", "DEX price\n(USD)",
               "Market cap\n(USD)"]
    FMTS = ["0", "@", "0.0###", BNB_FMT, BNB_FMT, USD_FMT, USD_FMT, "#,##0.00"]
    TABLE_TOP = 5

    def table(rows, col0):
        cols = list(range(col0, col0 + len(HEADERS)))
        styled_header(ws, TABLE_TOP, cols, HEADERS)
        for i, (step, bnb_in, p_bnb, p_usd, mcap, phase) in enumerate(rows):
            r = TABLE_TOP + 1 + i
            curve = phase == "curve"
            vals = [step, phase, round(bnb_in, 4),
                    round(p_bnb, 14) if curve else None,
                    None if curve else round(p_bnb, 14),
                    round(p_usd, 11) if curve else None,
                    None if curve else round(p_usd, 11),
                    round(mcap, 2) if mcap > 0 else None]
            for col, val, fmt in zip(cols, vals, FMTS):
                c = ws.cell(row=r, column=col, value=val)
                c.font = Font(size=9)
                c.number_format = fmt
                if i % 2:
                    c.fill = PatternFill("solid", fgColor=BAND)
        return TABLE_TOP + len(rows)

    last2 = table(v2, 6)   # F..M
    last3 = table(v3, 14)  # N..U

    # ---- Charts ------------------------------------------------------------
    def charts_for(label, col0, last_row, anchor_col, tgt_col):
        cats = Reference(ws, min_col=col0 + 2, min_row=TABLE_TOP + 1, max_row=last_row)
        y = TABLE_TOP

        def place(ch, offset):
            ch.set_categories(cats)
            ws.add_chart(ch, f"{get_column_letter(anchor_col)}{y + offset}")

        ch = make_chart(f"Token price in BNB — {label}", "Price (BNB per token)")
        ch.y_axis.number_format = BNB_FMT
        ch.series.append(line_series(ws, col0 + 3, TABLE_TOP + 1, last_row, "Curve price (BNB)", BLUE))
        ch.series.append(line_series(ws, col0 + 4, TABLE_TOP + 1, last_row, "DEX price (BNB)", ORANGE, marker=True))
        place(ch, 0)

        ch = make_chart(f"Token price in USD — {label}", "Price (USD per token)")
        ch.y_axis.number_format = USD_FMT
        ch.series.append(line_series(ws, col0 + 5, TABLE_TOP + 1, last_row, "Curve price (USD)", BLUE))
        ch.series.append(line_series(ws, col0 + 6, TABLE_TOP + 1, last_row, "DEX price (USD)", ORANGE, marker=True))
        place(ch, 21)

        ch = make_chart(f"Market cap vs $30k graduation target — {label}", "Market cap (USD)")
        ch.series.append(line_series(ws, col0 + 7, TABLE_TOP + 1, last_row, "Market cap (USD)", GREEN))
        ws.cell(row=TABLE_TOP, column=tgt_col, value="Graduation target")
        for r in range(TABLE_TOP + 1, last_row + 1):
            ws.cell(row=r, column=tgt_col, value=30000)
        ws.column_dimensions[get_column_letter(tgt_col)].hidden = True
        ch.series.append(line_series(ws, tgt_col, TABLE_TOP + 1, last_row, "Graduation target ($30k)", GRAY, width=16000, dashed=True))
        place(ch, 42)

    charts_for("PancakeSwap V2", 6, last2, 24, 22)   # charts at X, target col V
    charts_for("PancakeSwap V3", 14, last3, 37, 23)  # charts at AK, target col W

    # ---- Column widths -----------------------------------------------------
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    for col in "CD":
        ws.column_dimensions[col].width = 16
    for col0 in (6, 14):
        for i, w in enumerate([6, 7, 10, 17, 17, 15, 15, 12]):
            ws.column_dimensions[get_column_letter(col0 + i)].width = w

    wb.save(OUT)
    _scrub_scientific(OUT)


def _scrub_scientific(path):
    """Rewrite e-notation cell values to plain decimals — some viewers parse
    only the mantissa of '7.7e-09' and plot garbage."""
    def dec(m):
        return f"<v>{format(float(m.group(1)), '.20f').rstrip('0').rstrip('.')}</v>"
    zin = zipfile.ZipFile(path)
    items = [(i, zin.read(i.filename)) for i in zin.infolist()]
    zin.close()
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for info, data in items:
            if info.filename.startswith("xl/worksheets/"):
                data = re.sub(rb"<v>([0-9.]+[eE]-?\d+)</v>",
                              lambda m: dec(re.match(r"<v>(.*)</v>", m.group(0).decode())).encode(),
                              data)
            zout.writestr(info, data)


def main():
    if len(sys.argv) > 1:
        log = Path(sys.argv[1]).read_text()
    else:
        run = subprocess.run(
            ["forge", "test", "--match-contract", "GraduationCurveTest", "-vv"],
            cwd=ROOT / "foundry", capture_output=True, text=True, timeout=1800,
        )
        log = run.stdout
        if run.returncode != 0:
            # fork RPC flakiness is common — show why before bailing
            sys.exit(f"forge test failed:\n{run.stdout[-2000:]}{run.stderr[-500:]}")
    data = capture_rows(log)
    assert {"v2", "v3"} <= set(data), \
        f"markets found: {list(data)} — a capture test produced no rows (transient fork RPC? rerun)"
    build(data)
    print(f"wrote {OUT}  (v2: {len(data['v2'])} rows, v3: {len(data['v3'])} rows)")


if __name__ == "__main__":
    main()
